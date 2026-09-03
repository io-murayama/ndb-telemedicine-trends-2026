"""Input parsers for the NDB and official contextual datasets."""

from __future__ import annotations

import json
import math
import re
from pathlib import Path
from typing import Literal

from bs4 import BeautifulSoup
from openpyxl import load_workbook

PRIMARY_ONLINE_CODES = {"111014210", "112024210", "112024710"}
STANDARD_BASE_CODES = {"111000110", "112007410", "112011310"}
ONLINE_CODE_LABELS = {
    "111014210": "情報通信機器を用いた初診料",
    "112024210": "情報通信機器を用いた再診料",
    "112024710": "情報通信機器を用いた外来診療料",
}


def parse_number(value: object) -> float | None:
    """Return an NDB numeric cell, retaining masked cells as missing."""
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return None
    text = str(value).strip()
    if not text or text in {"-", "‐", "—"}:
        return None
    return float(text.replace(",", ""))


def procedure_code(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).replace(".0", "").strip()
    return text if re.fullmatch(r"\d{9}", text) else None


def _prefecture_locations(rows: list[tuple[object, ...]]) -> list[dict[str, object]]:
    codes, names = rows[2], rows[3]
    locations = []
    for column in range(6, len(codes)):
        code = str(codes[column]).zfill(2) if codes[column] is not None else ""
        name = str(names[column]).strip() if names[column] is not None else ""
        if re.fullmatch(r"\d{2}", code) and name:
            locations.append(
                {
                    "area_type": "prefecture",
                    "prefecture_code": code,
                    "sma_code": None,
                    "prefecture_name": name,
                    "area_name": name,
                    "column": column,
                }
            )
    if len(locations) != 47:
        raise ValueError(f"Expected 47 prefectures, found {len(locations)}")
    return locations


def _sma_locations(rows: list[tuple[object, ...]]) -> list[dict[str, object]]:
    codes, names = rows[4], rows[5]
    locations = []
    for column in range(6, len(codes)):
        code = str(codes[column]).strip() if codes[column] is not None else ""
        name = str(names[column]).strip() if names[column] is not None else ""
        if re.fullmatch(r"\d{4}", code) and name:
            locations.append(
                {
                    "area_type": "secondary_medical_area",
                    "prefecture_code": code[:2],
                    "sma_code": code,
                    "prefecture_name": None,
                    "area_name": name,
                    "column": column,
                }
            )
    if len(locations) < 300:
        raise ValueError(f"Expected secondary medical areas, found {len(locations)}")
    return locations


def read_ndb_geography(
    path: Path,
    fiscal_year: int,
    level: Literal["prefecture", "sma"],
) -> list[dict[str, object]]:
    """Extract three main online-care codes and standard basic-fee codes by area."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["外来"]
    rows = list(sheet.iter_rows(values_only=True))
    locations = _prefecture_locations(rows) if level == "prefecture" else _sma_locations(rows)
    data_start = 4 if level == "prefecture" else 6

    metrics = {
        (str(location["area_type"]), str(location["sma_code"] or location["prefecture_code"])): {
            **location,
            "fiscal_year": fiscal_year,
            "online_primary_count": 0.0,
            "online_primary_missing_components": 0,
            "standard_base_count": 0.0,
            "standard_base_missing_components": 0,
        }
        for location in locations
    }
    national = {
        "fiscal_year": fiscal_year,
        "online_primary_count": 0.0,
        "online_primary_missing_components": 0,
        "online_all_codes_count": 0.0,
        "online_all_codes_missing_components": 0,
        "standard_base_count": 0.0,
        "standard_base_missing_components": 0,
    }

    for row in rows[data_start:]:
        code = procedure_code(row[2] if len(row) > 2 else None)
        if code is None:
            continue
        name = str(row[3] or "")
        is_primary_online = code in PRIMARY_ONLINE_CODES
        is_standard_base = code in STANDARD_BASE_CODES
        is_any_online = "情報通信機器" in name
        if not (is_primary_online or is_standard_base or is_any_online):
            continue

        total = parse_number(row[5] if len(row) > 5 else None)
        if is_primary_online:
            _add(national, "online_primary", total)
        if is_standard_base:
            _add(national, "standard_base", total)
        if is_any_online:
            _add(national, "online_all_codes", total)

        for location in locations:
            key = (str(location["area_type"]), str(location["sma_code"] or location["prefecture_code"]))
            value = parse_number(row[int(location["column"])])
            if is_primary_online:
                _add(metrics[key], "online_primary", value)
            if is_standard_base:
                _add(metrics[key], "standard_base", value)

    return [*metrics.values(), {"national": national}]


def read_ndb_procedure_details(
    path: Path,
    fiscal_year: int,
    level: Literal["prefecture", "sma"],
    procedure_codes: set[str],
) -> list[dict[str, object]]:
    """Extract each selected procedure separately, retaining masked cells as missing.

    The NDB area tables are aggregated over procedures in the main analysis.  This
    narrower reader is used for the Shimane case study, where it is important to
    distinguish online initial visits from repeat visits without treating a masked
    value as zero.
    """
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["外来"]
    rows = list(sheet.iter_rows(values_only=True))
    locations = _prefecture_locations(rows) if level == "prefecture" else _sma_locations(rows)
    data_start = 4 if level == "prefecture" else 6
    records: list[dict[str, object]] = []

    for row in rows[data_start:]:
        code = procedure_code(row[2] if len(row) > 2 else None)
        if code not in procedure_codes:
            continue
        name = str(row[3] or "").strip()
        for location in locations:
            records.append(
                {
                    **location,
                    "fiscal_year": fiscal_year,
                    "procedure_code": code,
                    "procedure_name": name,
                    "count": parse_number(row[int(location["column"])]),
                }
            )

    expected = len(locations) * len(procedure_codes)
    if len(records) != expected:
        raise ValueError(f"Expected {expected} procedure-area records, found {len(records)}")
    return records


def _add(target: dict[str, object], prefix: str, value: float | None) -> None:
    count_key = f"{prefix}_count"
    missing_key = f"{prefix}_missing_components"
    if value is None:
        target[missing_key] = int(target.get(missing_key, 0)) + 1
    else:
        target[count_key] = float(target.get(count_key, 0.0)) + value


def read_population_age3(path: Path) -> list[dict[str, object]]:
    """Read 2021 population estimates by prefecture and broad age group."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["第3表"]
    records = []
    for row in sheet.iter_rows(min_row=13, values_only=True):
        code = str(row[1]).zfill(2) if row[1] is not None else ""
        name = str(row[2] or "").strip()
        if not re.fullmatch(r"\d{2}", code) or not name:
            continue
        under15, working_age, over65, over75 = (parse_number(row[i]) for i in (4, 5, 6, 7))
        if None in {under15, working_age, over65, over75}:
            continue
        population_known_thousand = under15 + working_age + over65
        records.append(
            {
                "prefecture_code": code,
                "prefecture_name": name,
                "population_known_thousand": population_known_thousand,
                "population_65plus_thousand": over65,
                "population_75plus_thousand": over75,
                "share_65plus": over65 / population_known_thousand,
                "share_75plus": over75 / population_known_thousand,
            }
        )
    if len(records) != 47:
        raise ValueError(f"Expected 47 population records, found {len(records)}")
    return records


def read_medical_facility_table105(path: Path) -> list[dict[str, object]]:
    """Read clinic counts from e-Stat's 2023 Medical Facility Survey table 105."""
    result = json.loads(path.read_text(encoding="utf-8"))
    soup = BeautifulSoup(result["table"], "lxml")
    table = soup.find("table", class_="stat-dbview-display-table-view")
    if table is None:
        raise ValueError("Could not find e-Stat table 105")
    records = []
    for row in table.select("tbody tr"):
        header = row.find("th")
        if header is None:
            continue
        source_code = str(header.get("data-unique", ""))
        name = header.get_text(" ", strip=True)
        if not re.fullmatch(r"00\d{3}", source_code) or name.startswith("（再掲）"):
            continue
        # e-Stat's location codes run from 00110 (Hokkaido) to 00570
        # (Okinawa), in steps of ten; they are not JIS prefecture codes.
        code = f"{int(source_code) // 10 - 10:02d}"
        if not (1 <= int(code) <= 47):
            continue
        cells = [parse_number(cell.get_text(" ", strip=True)) for cell in row.find_all("td")]
        if not cells or cells[0] is None:
            continue
        records.append(
            {
                "prefecture_code": code,
                "prefecture_name": name,
                "clinic_count_2023": cells[0],
                "remote_home_support_facilities_2023": cells[11] if len(cells) > 11 else None,
            }
        )
    if len(records) != 47:
        raise ValueError(f"Expected 47 medical-facility records, found {len(records)}")
    return records
